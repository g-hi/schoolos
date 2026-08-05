from __future__ import annotations

import io
import uuid
import zipfile
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from services.gateway.main import app
from shared.auth.jwt import get_current_user
from shared.auth.tenant import resolve_tenant
from shared.db.connection import get_db


def _tenant() -> SimpleNamespace:
    return SimpleNamespace(id=uuid.uuid4(), slug="greenwood", name="Greenwood", is_active=True, settings={})


def _actor(tenant_id: uuid.UUID, role: str = "principal") -> SimpleNamespace:
    return SimpleNamespace(id=uuid.uuid4(), tenant_id=tenant_id, role=role, is_active=True, name="Leader", email="leader@example.test")


def _db() -> AsyncMock:
    db = AsyncMock()
    db.add = MagicMock()
    db.commit = AsyncMock()
    return db


def _client(db: AsyncMock, tenant: SimpleNamespace, actor: SimpleNamespace) -> TestClient:
    app.dependency_overrides[get_db] = lambda: db
    app.dependency_overrides[resolve_tenant] = lambda: tenant
    app.dependency_overrides[get_current_user] = lambda: actor
    return TestClient(app, raise_server_exceptions=False)


def test_template_download_is_valid_xlsx_with_required_sheets() -> None:
    tenant = _tenant()
    actor = _actor(tenant.id)
    db = _db()

    with _client(db, tenant, actor) as client:
        response = client.get("/leadership/timetable-setup/imports/template")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    wb = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True, keep_vba=False)
    required = {
        "Instructions",
        "Teachers",
        "Classes",
        "Subjects",
        "Rooms",
        "School Week",
        "Periods",
        "Teaching Requirements",
        "Teacher Availability",
        "Fixed Sessions",
        "Constraints",
    }
    assert required.issubset(set(wb.sheetnames))
    wb.close()
    app.dependency_overrides.clear()


def test_template_has_no_macros_or_embedded_secrets() -> None:
    tenant = _tenant()
    actor = _actor(tenant.id)
    db = _db()

    with _client(db, tenant, actor) as client:
        response = client.get("/leadership/timetable-setup/imports/template")

    assert response.status_code == 200
    with zipfile.ZipFile(io.BytesIO(response.content)) as zf:
        names = set(zf.namelist())
    assert "xl/vbaProject.bin" not in names
    content_text = response.content.decode("latin-1", errors="ignore").lower()
    assert "database_url" not in content_text
    assert "secret" not in content_text
    app.dependency_overrides.clear()
