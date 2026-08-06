from __future__ import annotations

import hashlib
import io
import re
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import PurePath
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from services.gateway.ai.audit import log_action
from services.gateway.timetable_setup.readiness import compute_timetable_input_readiness
from shared.auth.dependencies import resolve_authenticated_leadership
from shared.auth.tenant import resolve_tenant
from shared.config import settings
from shared.db.connection import get_db, set_tenant_context
from shared.db.models import (
    AcademicYear,
    BellSchedule,
    BellSchedulePeriod,
    Campus,
    Class,
    ImportBatch,
    ImportRowResult,
    SchoolWeekConfig,
    Subject,
    Teacher,
    TeachingRoom,
    Tenant,
    Term,
    User,
    WeeklyTeachingRequirement,
)

router = APIRouter(prefix="/leadership/timetable-setup/imports", tags=["Timetable Workbook Imports"])

SUPPORTED_EXTENSION = ".xlsx"
REJECTED_EXTENSIONS = {".xls", ".xlsm"}
REQUIRED_SHEETS = {"teachers", "classes", "subjects", "periods", "teaching_requirements"}

SHEET_ALIASES: dict[str, tuple[str, ...]] = {
    "teachers": ("teachers", "staff", "teaching staff", "faculty"),
    "classes": ("classes", "class list", "sections", "homerooms"),
    "subjects": ("subjects", "courses", "curriculum"),
    "rooms": ("rooms", "classrooms", "facilities"),
    "school_week": ("school week", "school_week", "week config", "weekdays"),
    "periods": ("periods", "bell schedule", "timings", "school day"),
    "teaching_requirements": (
        "teaching requirements",
        "lesson allocation",
        "sessions",
        "subject allocation",
        "class subject requirements",
    ),
    "teacher_availability": ("teacher availability", "availability"),
    "fixed_sessions": ("fixed sessions", "fixed"),
    "constraints": ("constraints", "rules"),
    "instructions": ("instructions",),
}

CANONICAL_SHEETS = [
    "Instructions",
    "Teachers",
    "Classes",
    "Subjects",
    "Rooms",
    "School Week",
    "Periods",
    "Teaching Requirements",
    "Teacher Availability",
    "Fixed Sessions",
    "Constraints",
]

CANONICAL_COLUMNS: dict[str, list[str]] = {
    "teachers": [
        "teacher_id",
        "teacher_name",
        "email",
        "department",
        "subjects",
        "campus_code",
        "maximum_weekly_load",
        "active",
    ],
    "classes": ["class_code", "class_name", "grade_level", "campus_code", "home_room_code", "active"],
    "subjects": ["subject_code", "subject_name", "department", "specialist_room_type", "active"],
    "rooms": [
        "room_code",
        "room_name",
        "campus_code",
        "room_type",
        "capacity",
        "specialist_capabilities",
        "active",
    ],
    "school_week": ["campus_code", "weekday", "operational", "default_schedule_name"],
    "periods": [
        "schedule_name",
        "campus_code",
        "period_number",
        "period_label",
        "start_time",
        "end_time",
        "period_type",
        "teaching_period",
        "effective_from",
        "effective_to",
    ],
    "teaching_requirements": [
        "academic_year",
        "term",
        "campus_code",
        "class_code",
        "subject_code",
        "teacher_id",
        "sessions_per_week",
        "periods_per_session",
        "minimum_daily_sessions",
        "maximum_daily_sessions",
        "double_period_preference",
        "required_room_type",
        "preferred_periods",
        "forbidden_periods",
        "priority",
    ],
    "teacher_availability": ["teacher_id", "weekday", "period_number", "availability", "reason"],
    "fixed_sessions": ["class_code", "subject_code", "teacher_id", "weekday", "period_number", "room_code", "fixed"],
    "constraints": ["constraint_type", "target_type", "target_identifier", "value", "priority", "hard_or_soft", "explanation"],
}

REQUIRED_FIELDS: dict[str, set[str]] = {
    "teachers": {"teacher_id", "teacher_name", "active"},
    "classes": {"class_code", "class_name", "active"},
    "subjects": {"subject_code", "subject_name", "active"},
    "rooms": {"room_code", "room_name", "room_type", "active"},
    "school_week": {"weekday", "operational"},
    "periods": {"schedule_name", "period_number", "start_time", "end_time", "teaching_period"},
    "teaching_requirements": {"academic_year", "term", "class_code", "subject_code", "sessions_per_week"},
}

COLUMN_ALIASES: dict[str, str] = {
    "teacher name": "teacher_name",
    "staff name": "teacher_name",
    "employee id": "teacher_id",
    "lessons per week": "sessions_per_week",
    "periods weekly": "sessions_per_week",
    "class": "class_code",
    "section": "class_code",
    "subject": "subject_code",
    "room": "room_code",
    "start": "start_time",
    "end": "end_time",
    "code": "subject_code",
}

TIME_RE = re.compile(r"^([01]\d|2[0-3]):[0-5]\d$")


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _sanitize_filename(filename: str | None) -> str:
    base = PurePath(filename or "workbook.xlsx").name.replace("\x00", "")
    return (base[:255] or "workbook.xlsx").strip()


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _norm(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", _norm(value)).strip("_")


def _parse_bool(value: str) -> bool | None:
    v = _norm(value)
    if v in {"true", "yes", "y", "1", "active", "enabled"}:
        return True
    if v in {"false", "no", "n", "0", "inactive", "disabled"}:
        return False
    return None


def _parse_int(value: str) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _parse_period_list(value: str) -> list[int]:
    if not value:
        return []
    out: list[int] = []
    for token in value.split(","):
        token = token.strip()
        if not token:
            continue
        num = _parse_int(token)
        if num is not None:
            out.append(num)
    return out


def _template_instructions() -> list[list[str]]:
    return [
        ["SchoolOS Timetable Workbook Instructions", ""],
        ["Purpose", "Upload for preview, deterministic mapping, validation, leadership approval, and controlled commit."],
        ["File Type", "Only .xlsx is supported. .xls and .xlsm are rejected."],
        ["Required Sheets", "Teachers, Classes, Subjects, Periods, Teaching Requirements"],
        ["Optional Sheets", "Rooms, School Week, Teacher Availability, Fixed Sessions, Constraints"],
        ["Cross Sheet References", "Use stable IDs/codes: class_code, subject_code, teacher_id, room_code, schedule_name, campus_code."],
        ["Date Format", "YYYY-MM-DD"],
        ["Time Format", "HH:MM 24-hour"],
        ["Lifecycle", "Upload -> mapping review -> validate -> leadership commit"],
        ["Policy", "Uploaded data does not become operational until leadership commit completes successfully."],
        ["Notes", "Keep headers in the first row where possible. Optional sheets can be left empty."],
    ]


def _build_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    for row in _template_instructions():
        ws.append(row)

    sheet_to_entity = {
        "Teachers": "teachers",
        "Classes": "classes",
        "Subjects": "subjects",
        "Rooms": "rooms",
        "School Week": "school_week",
        "Periods": "periods",
        "Teaching Requirements": "teaching_requirements",
        "Teacher Availability": "teacher_availability",
        "Fixed Sessions": "fixed_sessions",
        "Constraints": "constraints",
    }
    for sheet in CANONICAL_SHEETS[1:]:
        s = wb.create_sheet(sheet)
        entity = sheet_to_entity[sheet]
        s.append(CANONICAL_COLUMNS[entity])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


def _ensure_actor_tenant(actor: User, tenant: Tenant) -> None:
    if not actor.is_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Inactive users cannot access this resource.")
    if actor.tenant_id != tenant.id:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid or expired token.")


def _detect_entity(sheet_name: str, headers: list[str]) -> tuple[str | None, float, str]:
    name = _norm(sheet_name)
    for entity, aliases in SHEET_ALIASES.items():
        if name in aliases:
            confidence = 1.0 if name == entity.replace("_", " ") else 0.9
            return entity, confidence, f"matched alias '{name}'"

    best_entity: str | None = None
    best_score = 0.0
    for entity, columns in CANONICAL_COLUMNS.items():
        if entity == "instructions":
            continue
        expected = set(columns)
        observed = set(_slug(value) for value in headers if value)
        if not expected:
            continue
        score = len(expected & observed) / float(len(expected))
        if score > best_score:
            best_score = score
            best_entity = entity
    if best_entity and best_score >= 0.4:
        return best_entity, round(best_score, 2), "header similarity"
    return None, 0.0, "no deterministic alias or header match"


def _propose_column_mappings(entity: str, headers: list[str], sample_rows: list[dict[str, str]]) -> tuple[list[dict[str, Any]], list[str]]:
    mappings: list[dict[str, Any]] = []
    used_targets: set[str] = set()
    canonical = set(CANONICAL_COLUMNS.get(entity, []))

    for header in headers:
        src = _clean(header)
        source_key = _slug(src)
        target: str | None = None
        confidence = 0.0
        reason = ""
        if source_key in canonical:
            target = source_key
            confidence = 1.0
            reason = "exact canonical header"
        elif _norm(src) in COLUMN_ALIASES:
            candidate = COLUMN_ALIASES[_norm(src)]
            if candidate in canonical:
                target = candidate
                confidence = 0.9
                reason = "known alias"
        elif source_key.endswith("_id") and source_key in canonical:
            target = source_key
            confidence = 0.8
            reason = "identifier pattern"

        if target and target in used_targets:
            target = None
            confidence = 0.0
            reason = "target already mapped by another source column"
        if target:
            used_targets.add(target)

        samples = []
        for row in sample_rows[:5]:
            value = _clean(row.get(src) or row.get(source_key) or "")
            if value:
                samples.append(value)
            if len(samples) >= 3:
                break

        mappings.append(
            {
                "source_column": src,
                "target_field": target,
                "confidence": confidence,
                "reason": reason or "unmapped",
                "required": bool(target and target in REQUIRED_FIELDS.get(entity, set())),
                "data_type_compatible": True,
                "sample_values": samples,
                "confirmed": bool(target and confidence >= 0.9),
            }
        )

    mapped_targets = {item["target_field"] for item in mappings if item.get("target_field")}
    required_unmapped = sorted(REQUIRED_FIELDS.get(entity, set()) - mapped_targets)
    return mappings, required_unmapped


def _canonical_row_from_mapping(row: dict[str, str], mappings: list[dict[str, Any]]) -> dict[str, str]:
    out: dict[str, str] = {}
    for mapping in mappings:
        target = mapping.get("target_field")
        if not target:
            continue
        source = mapping.get("source_column")
        if not source:
            continue
        value = _clean(row.get(source))
        out[target] = value
    return out


def _workbook_zip_guard(content: bytes, filename: str) -> None:
    if len(content) <= 0:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Uploaded workbook is empty.")
    if len(content) > settings.timetable_workbook_max_upload_bytes:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Workbook exceeds upload size limit.")

    extension = PurePath(filename).suffix.lower()
    if extension in REJECTED_EXTENSIONS:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Unsupported workbook extension: {extension}.")
    if extension != SUPPORTED_EXTENSION:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Only .xlsx workbook uploads are supported.")

    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            if zf.testzip() is not None:
                raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Malformed XLSX archive.")
            names = {name for name in zf.namelist()}
            if "xl/vbaProject.bin" in names:
                raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Macro-enabled workbooks are not supported.")
            if any(name.startswith("xl/externalLinks/") for name in names):
                raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="External workbook links are not supported.")
            if "[Content_Types].xml" not in names:
                raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Malformed XLSX archive.")
    except zipfile.BadZipFile:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid or password-protected XLSX file.") from None


async def _load_batch(db: AsyncSession, tenant_id: uuid.UUID, batch_id: uuid.UUID, *, lock: bool = False) -> ImportBatch | None:
    stmt = select(ImportBatch).where(ImportBatch.id == batch_id, ImportBatch.tenant_id == tenant_id, ImportBatch.entity_type == "timetable_workbook")
    if lock:
        stmt = stmt.with_for_update()
    return await db.scalar(stmt)


def _batch_payload(batch: ImportBatch) -> dict[str, Any]:
    meta = batch.metadata_json or {}
    return {
        "id": str(batch.id),
        "tenant_id": str(batch.tenant_id),
        "entity_type": batch.entity_type,
        "original_filename": batch.original_filename,
        "file_sha256": batch.file_sha256,
        "status": batch.status,
        "mode": batch.mode,
        "import_format": batch.import_format,
        "created_by_user_id": str(batch.created_by_user_id),
        "total_rows": batch.total_rows,
        "valid_rows": batch.valid_rows,
        "invalid_rows": batch.invalid_rows,
        "created_rows": batch.created_rows,
        "updated_rows": batch.updated_rows,
        "skipped_rows": batch.skipped_rows,
        "conflict_rows": batch.conflict_rows,
        "started_at": batch.started_at,
        "completed_at": batch.completed_at,
        "committed_at": batch.committed_at,
        "created_at": batch.created_at,
        "updated_at": batch.updated_at,
        "workbook_summary": meta.get("workbook", {}),
    }


def _list_to_page(items: list[dict[str, Any]], page: int, page_size: int) -> dict[str, Any]:
    total = len(items)
    start = max(0, (page - 1) * page_size)
    return {"items": items[start : start + page_size], "total": total, "page": page, "page_size": page_size}


@router.get("/template", summary="Download official SchoolOS timetable workbook template")
async def download_template(
    tenant: Tenant = Depends(resolve_tenant),
    actor: User = Depends(resolve_authenticated_leadership),
    db: AsyncSession = Depends(get_db),
):
    _ensure_actor_tenant(actor, tenant)
    await set_tenant_context(db, tenant.id)
    content = _build_template_bytes()
    await log_action(
        db=db,
        tenant_id=tenant.id,
        action="timetable_setup.workbook.template_downloaded",
        entity_type="ImportBatch",
        entity_id=None,
        actor_id=actor.id,
        details={"format": "xlsx"},
    )
    await db.commit()
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="schoolos_timetable_template.xlsx"'},
    )


@router.post("/workbooks", summary="Upload timetable workbook for mapping and preview")
async def upload_workbook(
    file: UploadFile = File(...),
    tenant: Tenant = Depends(resolve_tenant),
    actor: User = Depends(resolve_authenticated_leadership),
    db: AsyncSession = Depends(get_db),
):
    _ensure_actor_tenant(actor, tenant)
    await set_tenant_context(db, tenant.id)

    safe_name = _sanitize_filename(file.filename)
    content = await file.read()
    _workbook_zip_guard(content, safe_name)

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True, keep_vba=False, keep_links=False)
    except Exception:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Failed to parse XLSX workbook.") from None

    if len(workbook.sheetnames) > settings.timetable_workbook_max_sheets:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Workbook sheet count exceeds limit.")

    meta_sheets: list[dict[str, Any]] = []
    row_models: list[ImportRowResult] = []
    global_row_number = 1
    unresolved_mapping_count = 0

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        if ws.max_row > settings.timetable_workbook_max_rows_per_sheet:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Sheet '{sheet_name}' exceeds row limit.")
        if ws.max_column > settings.timetable_workbook_max_columns_per_sheet:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Sheet '{sheet_name}' exceeds column limit.")

        header_row_num = 1
        headers = [_clean(cell) for cell in next(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True), ())]
        headers = [value for value in headers if value]
        entity, confidence, reason = _detect_entity(sheet_name, headers)

        sample_rows: list[dict[str, str]] = []
        data_rows: list[dict[str, str]] = []
        if headers:
            for values in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
                row = {headers[idx]: _clean(values[idx]) if idx < len(values) else "" for idx in range(len(headers))}
                if not any(row.values()):
                    continue
                data_rows.append(row)
                if len(sample_rows) < settings.timetable_workbook_preview_sample_rows:
                    sample_rows.append(row)

        mappings: list[dict[str, Any]] = []
        required_unmapped: list[str] = []
        if entity and entity != "instructions":
            mappings, required_unmapped = _propose_column_mappings(entity, headers, sample_rows)
            unresolved_mapping_count += len(required_unmapped)

        review_status = "confirmed" if confidence >= 0.9 and not required_unmapped else "needs_review"
        ignored = entity == "instructions"

        meta_sheets.append(
            {
                "original_sheet_name": sheet_name,
                "proposed_entity_type": entity,
                "confidence": confidence,
                "reason": reason,
                "row_count": len(data_rows),
                "header_row_number": header_row_num,
                "review_status": review_status,
                "ignored": ignored,
                "mappings": mappings,
                "required_unmapped_fields": required_unmapped,
                "sample_rows": sample_rows,
            }
        )

        for idx, row in enumerate(data_rows, start=1):
            row_models.append(
                ImportRowResult(
                    id=uuid.uuid4(),
                    tenant_id=tenant.id,
                    import_batch_id=uuid.uuid4(),
                    row_number=global_row_number,
                    status="valid",
                    action="none",
                    entity_reference_id=None,
                    error_code=None,
                    error_message=None,
                    severity=None,
                    sheet_name=sheet_name,
                    source_column=None,
                    field_name=None,
                    field_errors={"record_type": "row_data", "sheet_row_number": idx},
                    normalized_data={},
                    row_data=row,
                    created_at=_now(),
                    updated_at=_now(),
                )
            )
            global_row_number += 1

    workbook.close()

    status_value = "mapping_required" if unresolved_mapping_count > 0 else "preview_ready"
    batch_id = uuid.uuid4()
    for row_model in row_models:
        row_model.import_batch_id = batch_id

    batch = ImportBatch(
        id=batch_id,
        tenant_id=tenant.id,
        entity_type="timetable_workbook",
        original_filename=safe_name,
        file_sha256=hashlib.sha256(content).hexdigest(),
        status=status_value,
        mode="workbook",
        import_format="xlsx",
        created_by_user_id=actor.id,
        total_rows=len(row_models),
        valid_rows=len(row_models),
        invalid_rows=0,
        created_rows=0,
        updated_rows=0,
        skipped_rows=0,
        conflict_rows=0,
        started_at=_now(),
        completed_at=_now(),
        committed_at=None,
        metadata_json={
            "version": 1,
            "workbook": {
                "filename": safe_name,
                "sheet_count": len(meta_sheets),
                "all_data_from_official_template": all(_norm(item["original_sheet_name"]) in {_norm(s) for s in CANONICAL_SHEETS} for item in meta_sheets),
            },
            "workflow": {"required_entities": sorted(REQUIRED_SHEETS)},
            "sheets": meta_sheets,
        },
    )
    db.add(batch)
    for row_model in row_models:
        db.add(row_model)

    await log_action(
        db=db,
        tenant_id=tenant.id,
        action="timetable_setup.workbook.uploaded",
        entity_type="ImportBatch",
        entity_id=batch.id,
        actor_id=actor.id,
        details={"filename": safe_name, "sha256": batch.file_sha256, "sheet_count": len(meta_sheets)},
    )
    await db.commit()
    await db.refresh(batch)

    return {
        "batch": _batch_payload(batch),
        "detected_sheets": batch.metadata_json.get("sheets", []),
        "unresolved_mapping_count": unresolved_mapping_count,
    }


@router.get("/workbooks", summary="List timetable workbook batches")
async def list_workbooks(
    status_filter: str | None = Query(default=None, alias="status"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    tenant: Tenant = Depends(resolve_tenant),
    actor: User = Depends(resolve_authenticated_leadership),
    db: AsyncSession = Depends(get_db),
):
    _ensure_actor_tenant(actor, tenant)
    await set_tenant_context(db, tenant.id)

    stmt = select(ImportBatch).where(ImportBatch.tenant_id == tenant.id, ImportBatch.entity_type == "timetable_workbook").order_by(ImportBatch.created_at.desc())
    if status_filter:
        stmt = stmt.where(ImportBatch.status == status_filter)
    rows = (await db.execute(stmt)).scalars().all()
    payload = [_batch_payload(item) for item in rows]
    return _list_to_page(payload, page, page_size)


@router.get("/workbooks/{batch_id}", summary="Get one workbook batch")
async def get_workbook_batch(
    batch_id: uuid.UUID,
    tenant: Tenant = Depends(resolve_tenant),
    actor: User = Depends(resolve_authenticated_leadership),
    db: AsyncSession = Depends(get_db),
):
    _ensure_actor_tenant(actor, tenant)
    await set_tenant_context(db, tenant.id)
    batch = await _load_batch(db, tenant.id, batch_id)
    if batch is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Workbook batch not found.")
    return _batch_payload(batch)


@router.get("/workbooks/{batch_id}/sheets", summary="Get detected workbook sheets and mappings")
async def get_workbook_sheets(
    batch_id: uuid.UUID,
    tenant: Tenant = Depends(resolve_tenant),
    actor: User = Depends(resolve_authenticated_leadership),
    db: AsyncSession = Depends(get_db),
):
    _ensure_actor_tenant(actor, tenant)
    await set_tenant_context(db, tenant.id)
    batch = await _load_batch(db, tenant.id, batch_id)
    if batch is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Workbook batch not found.")
    return {
        "batch_id": str(batch.id),
        "status": batch.status,
        "sheets": (batch.metadata_json or {}).get("sheets", []),
    }


@router.get("/workbooks/{batch_id}/preview", summary="Get paginated row preview")
async def get_workbook_preview(
    batch_id: uuid.UUID,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=200),
    tenant: Tenant = Depends(resolve_tenant),
    actor: User = Depends(resolve_authenticated_leadership),
    db: AsyncSession = Depends(get_db),
):
    _ensure_actor_tenant(actor, tenant)
    await set_tenant_context(db, tenant.id)
    batch = await _load_batch(db, tenant.id, batch_id)
    if batch is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Workbook batch not found.")

    row_models = (
        await db.execute(
            select(ImportRowResult)
            .where(
                ImportRowResult.import_batch_id == batch.id,
                ImportRowResult.tenant_id == tenant.id,
                ImportRowResult.severity.is_(None),
            )
            .order_by(ImportRowResult.row_number.asc())
        )
    ).scalars().all()
    items = [
        {
            "row_number": row.row_number,
            "sheet_name": row.sheet_name,
            "row_data": row.row_data,
            "normalized_data": row.normalized_data,
        }
        for row in row_models
    ]

    diagnostics = (
        await db.execute(
            select(ImportRowResult)
            .where(
                ImportRowResult.import_batch_id == batch.id,
                ImportRowResult.tenant_id == tenant.id,
                ImportRowResult.severity.is_not(None),
            )
        )
    ).scalars().all()
    blocker_count = sum(1 for d in diagnostics if d.severity == "blocker")
    warning_count = sum(1 for d in diagnostics if d.severity == "warning")
    information_count = sum(1 for d in diagnostics if d.severity == "information")

    unresolved_mapping = sum(len(sheet.get("required_unmapped_fields", [])) for sheet in (batch.metadata_json or {}).get("sheets", []))

    return {
        "batch": _batch_payload(batch),
        "preview_rows": _list_to_page(items, page, page_size),
        "counts": {
            "blocker_count": blocker_count,
            "warning_count": warning_count,
            "information_count": information_count,
            "unresolved_mapping_count": unresolved_mapping,
            "row_count": len(items),
        },
    }


@router.patch("/workbooks/{batch_id}/mappings", summary="Confirm or override sheet/column mappings")
async def update_workbook_mappings(
    batch_id: uuid.UUID,
    body: dict[str, Any],
    tenant: Tenant = Depends(resolve_tenant),
    actor: User = Depends(resolve_authenticated_leadership),
    db: AsyncSession = Depends(get_db),
):
    _ensure_actor_tenant(actor, tenant)
    await set_tenant_context(db, tenant.id)
    batch = await _load_batch(db, tenant.id, batch_id, lock=True)
    if batch is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Workbook batch not found.")
    if batch.status in {"committed", "cancelled"}:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Finished batches cannot be remapped.")

    incoming = body.get("sheets")
    if not isinstance(incoming, list):
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Body must include sheets list.")

    meta = batch.metadata_json or {}
    current_sheets = meta.get("sheets", [])
    by_name = {item.get("original_sheet_name"): item for item in current_sheets}

    for item in incoming:
        source_name = item.get("sheet_name")
        if source_name not in by_name:
            continue
        target = by_name[source_name]
        if "ignored" in item:
            target["ignored"] = bool(item.get("ignored"))
        if "entity_type" in item and item.get("entity_type"):
            target["proposed_entity_type"] = _slug(str(item.get("entity_type")))
            target["review_status"] = "confirmed"
        if "header_row_number" in item:
            header_row = int(item.get("header_row_number"))
            target["header_row_number"] = max(1, header_row)
        if "mappings" in item and isinstance(item["mappings"], dict):
            override_map: dict[str, str | None] = item["mappings"]
            for mapping in target.get("mappings", []):
                src = mapping.get("source_column")
                if src in override_map:
                    desired = override_map[src]
                    mapping["target_field"] = _slug(str(desired)) if desired else None
                    mapping["confirmed"] = True
                    mapping["confidence"] = 1.0 if desired else 0.0
                    mapping["reason"] = "administrator override" if desired else "administrator ignored"

            mapped_targets = {m.get("target_field") for m in target.get("mappings", []) if m.get("target_field")}
            required = REQUIRED_FIELDS.get(target.get("proposed_entity_type") or "", set())
            target["required_unmapped_fields"] = sorted(required - mapped_targets)

    unresolved = sum(len(sheet.get("required_unmapped_fields", [])) for sheet in current_sheets if not sheet.get("ignored"))
    batch.metadata_json = meta
    batch.status = "mapping_required" if unresolved > 0 else "preview_ready"
    batch.updated_at = _now()

    await log_action(
        db=db,
        tenant_id=tenant.id,
        action="timetable_setup.workbook.mappings_updated",
        entity_type="ImportBatch",
        entity_id=batch.id,
        actor_id=actor.id,
        details={"unresolved_mapping_count": unresolved},
    )
    await db.commit()
    await db.refresh(batch)
    return {"batch": _batch_payload(batch), "sheets": current_sheets, "unresolved_mapping_count": unresolved}


def _diag(
    *,
    sheet: str,
    row_number: int,
    severity: str,
    code: str,
    message: str,
    field: str | None = None,
    source_column: str | None = None,
    original_value: Any = None,
    normalized_value: Any = None,
    related_entity: str | None = None,
    recommendation: str | None = None,
) -> dict[str, Any]:
    return {
        "sheet_name": sheet,
        "row_number": row_number,
        "severity": severity,
        "code": code,
        "message": message,
        "field_name": field,
        "source_column": source_column,
        "original_value": original_value,
        "normalized_value": normalized_value,
        "related_entity": related_entity,
        "recommended_correction": recommendation,
    }


async def _validate_batch(batch: ImportBatch, tenant: Tenant, db: AsyncSession) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    meta = batch.metadata_json or {}
    sheets = meta.get("sheets", [])
    by_sheet_name = {item.get("original_sheet_name"): item for item in sheets}

    row_models = (
        await db.execute(
            select(ImportRowResult)
            .where(
                ImportRowResult.import_batch_id == batch.id,
                ImportRowResult.tenant_id == tenant.id,
                ImportRowResult.severity.is_(None),
            )
            .order_by(ImportRowResult.row_number.asc())
        )
    ).scalars().all()

    normalized_by_entity: dict[str, list[dict[str, Any]]] = {}
    diagnostics: list[dict[str, Any]] = []

    # Required sheet checks
    present_entities = {sheet.get("proposed_entity_type") for sheet in sheets if not sheet.get("ignored")}
    for required in REQUIRED_SHEETS:
        if required not in present_entities:
            diagnostics.append(
                _diag(
                    sheet="(workbook)",
                    row_number=1,
                    severity="blocker",
                    code="missing_required_sheet",
                    message=f"Missing required sheet for workflow: {required}.",
                    recommendation="Add the required sheet or map an existing sheet to this entity.",
                )
            )

    for sheet in sheets:
        if sheet.get("ignored"):
            diagnostics.append(
                _diag(
                    sheet=sheet.get("original_sheet_name") or "",
                    row_number=1,
                    severity="information",
                    code="optional_sheet_ignored",
                    message="Sheet is ignored by administrator choice.",
                )
            )
            continue

        entity = sheet.get("proposed_entity_type")
        if not entity:
            diagnostics.append(
                _diag(
                    sheet=sheet.get("original_sheet_name") or "",
                    row_number=1,
                    severity="blocker",
                    code="unresolved_sheet_mapping",
                    message="Sheet classification is unresolved.",
                )
            )
            continue

        unresolved = sheet.get("required_unmapped_fields", [])
        if unresolved:
            diagnostics.append(
                _diag(
                    sheet=sheet.get("original_sheet_name") or "",
                    row_number=1,
                    severity="blocker",
                    code="missing_required_column_mapping",
                    message=f"Required fields are unmapped: {', '.join(unresolved)}.",
                    recommendation="Confirm or override column mappings before validation.",
                )
            )

        mappings = sheet.get("mappings", [])
        entity_rows: list[dict[str, Any]] = []
        for row in row_models:
            if row.sheet_name != sheet.get("original_sheet_name"):
                continue
            canonical = _canonical_row_from_mapping(row.row_data or {}, mappings)
            canonical["__global_row"] = row.row_number
            canonical["__sheet_row"] = int((row.field_errors or {}).get("sheet_row_number") or 0)
            entity_rows.append(canonical)
        normalized_by_entity[entity] = entity_rows

    # Deterministic diagnostics
    for entity_key, code_field in (("teachers", "teacher_id"), ("classes", "class_code"), ("subjects", "subject_code"), ("rooms", "room_code")):
        seen: set[str] = set()
        for row in normalized_by_entity.get(entity_key, []):
            value = _norm(row.get(code_field, ""))
            if not value:
                diagnostics.append(
                    _diag(
                        sheet=entity_key,
                        row_number=row.get("__sheet_row", 0),
                        severity="blocker",
                        code="missing_required_field",
                        message=f"{code_field} is required.",
                        field=code_field,
                    )
                )
                continue
            if value in seen:
                diagnostics.append(
                    _diag(
                        sheet=entity_key,
                        row_number=row.get("__sheet_row", 0),
                        severity="blocker",
                        code=f"duplicate_{code_field}",
                        message=f"Duplicate value in workbook: {value}.",
                        field=code_field,
                    )
                )
            seen.add(value)

    # Period validity
    period_rows = normalized_by_entity.get("periods", [])
    schedule_buckets: dict[str, list[tuple[str, str, int]]] = {}
    for row in period_rows:
        start_time = _clean(row.get("start_time"))
        end_time = _clean(row.get("end_time"))
        if not TIME_RE.fullmatch(start_time) or not TIME_RE.fullmatch(end_time):
            diagnostics.append(
                _diag(
                    sheet="periods",
                    row_number=row.get("__sheet_row", 0),
                    severity="blocker",
                    code="invalid_time",
                    message="start_time and end_time must use HH:MM 24-hour format.",
                )
            )
            continue
        if start_time >= end_time:
            diagnostics.append(
                _diag(
                    sheet="periods",
                    row_number=row.get("__sheet_row", 0),
                    severity="blocker",
                    code="invalid_time_range",
                    message="start_time must be earlier than end_time.",
                )
            )
            continue
        schedule_name = _clean(row.get("schedule_name")) or "default"
        schedule_buckets.setdefault(schedule_name, []).append((start_time, end_time, row.get("__sheet_row", 0)))

    for schedule_name, segments in schedule_buckets.items():
        for index, item in enumerate(sorted(segments, key=lambda x: x[0])):
            if index == 0:
                continue
            prev = sorted(segments, key=lambda x: x[0])[index - 1]
            if item[0] < prev[1]:
                diagnostics.append(
                    _diag(
                        sheet="periods",
                        row_number=item[2],
                        severity="blocker",
                        code="overlapping_periods",
                        message=f"Overlapping period range in schedule '{schedule_name}'.",
                    )
                )

    # Teaching requirement references and numeric checks
    class_codes = {_norm(item.get("class_code", "")) for item in normalized_by_entity.get("classes", [])}
    subject_codes = {_norm(item.get("subject_code", "")) for item in normalized_by_entity.get("subjects", [])}
    teacher_ids = {_norm(item.get("teacher_id", "")) for item in normalized_by_entity.get("teachers", [])}

    for row in normalized_by_entity.get("teaching_requirements", []):
        class_code = _norm(_clean(row.get("class_code")))
        subject_code = _norm(_clean(row.get("subject_code")))
        teacher_id = _norm(_clean(row.get("teacher_id")))

        if class_code and class_code not in class_codes:
            existing_class = await db.scalar(select(Class.id).where(Class.tenant_id == tenant.id, func.lower(Class.code) == class_code, Class.is_active.is_(True)))
            if existing_class is None:
                diagnostics.append(_diag(sheet="teaching_requirements", row_number=row.get("__sheet_row", 0), severity="blocker", code="unknown_class", message=f"Unknown class_code '{class_code}'."))

        if subject_code and subject_code not in subject_codes:
            existing_subject = await db.scalar(select(Subject.id).where(Subject.tenant_id == tenant.id, func.lower(Subject.code) == subject_code))
            if existing_subject is None:
                diagnostics.append(_diag(sheet="teaching_requirements", row_number=row.get("__sheet_row", 0), severity="blocker", code="unknown_subject", message=f"Unknown subject_code '{subject_code}'."))

        if teacher_id and teacher_id not in teacher_ids:
            existing_teacher = await db.scalar(select(Teacher.id).where(Teacher.tenant_id == tenant.id, func.lower(Teacher.employee_id) == teacher_id))
            if existing_teacher is None:
                diagnostics.append(_diag(sheet="teaching_requirements", row_number=row.get("__sheet_row", 0), severity="blocker", code="unknown_teacher", message=f"Unknown teacher_id '{teacher_id}'."))

        sessions_per_week = _parse_int(_clean(row.get("sessions_per_week")))
        periods_per_session = _parse_int(_clean(row.get("periods_per_session")) or "1")
        if sessions_per_week is None or sessions_per_week <= 0:
            diagnostics.append(_diag(sheet="teaching_requirements", row_number=row.get("__sheet_row", 0), severity="blocker", code="invalid_sessions_per_week", message="sessions_per_week must be positive."))
        if periods_per_session is None or periods_per_session <= 0:
            diagnostics.append(_diag(sheet="teaching_requirements", row_number=row.get("__sheet_row", 0), severity="blocker", code="invalid_periods_per_session", message="periods_per_session must be positive."))
        if sessions_per_week and sessions_per_week > 60:
            diagnostics.append(_diag(sheet="teaching_requirements", row_number=row.get("__sheet_row", 0), severity="warning", code="high_sessions_per_week", message="Unusually high weekly session requirement."))

    # Optional warnings
    for row in normalized_by_entity.get("teachers", []):
        if not _clean(row.get("email")):
            diagnostics.append(_diag(sheet="teachers", row_number=row.get("__sheet_row", 0), severity="warning", code="teacher_email_absent", message="Teacher email is absent."))
        if not _clean(row.get("department")):
            diagnostics.append(_diag(sheet="teachers", row_number=row.get("__sheet_row", 0), severity="warning", code="teacher_department_absent", message="Teacher department is absent."))

    for row in normalized_by_entity.get("rooms", []):
        if not _clean(row.get("capacity")):
            diagnostics.append(_diag(sheet="rooms", row_number=row.get("__sheet_row", 0), severity="warning", code="room_capacity_absent", message="Room capacity is absent."))

    if meta.get("workbook", {}).get("all_data_from_official_template"):
        diagnostics.append(
            _diag(
                sheet="(workbook)",
                row_number=1,
                severity="information",
                code="official_template_detected",
                message="All data originated from the official SchoolOS template sheet names.",
            )
        )

    return diagnostics, normalized_by_entity


@router.post("/workbooks/{batch_id}/validate", summary="Validate workbook mappings and rows")
async def validate_workbook(
    batch_id: uuid.UUID,
    tenant: Tenant = Depends(resolve_tenant),
    actor: User = Depends(resolve_authenticated_leadership),
    db: AsyncSession = Depends(get_db),
):
    _ensure_actor_tenant(actor, tenant)
    await set_tenant_context(db, tenant.id)
    batch = await _load_batch(db, tenant.id, batch_id, lock=True)
    if batch is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Workbook batch not found.")
    if batch.status in {"cancelled", "committed"}:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Finished batches cannot be validated.")

    diagnostics, normalized_by_entity = await _validate_batch(batch, tenant, db)

    await db.execute(
        delete(ImportRowResult).where(
            ImportRowResult.import_batch_id == batch.id,
            ImportRowResult.tenant_id == tenant.id,
            ImportRowResult.severity.is_not(None),
        )
    )

    max_row_number = await db.scalar(
        select(func.max(ImportRowResult.row_number)).where(
            ImportRowResult.import_batch_id == batch.id,
            ImportRowResult.tenant_id == tenant.id,
        )
    )
    next_row = int(max_row_number or 0) + 1

    blockers = 0
    warnings = 0
    infos = 0
    for diag in diagnostics:
        severity = diag["severity"]
        if severity == "blocker":
            blockers += 1
        elif severity == "warning":
            warnings += 1
        else:
            infos += 1

        db.add(
            ImportRowResult(
                id=uuid.uuid4(),
                tenant_id=tenant.id,
                import_batch_id=batch.id,
                row_number=next_row,
                status="invalid" if severity == "blocker" else "valid",
                action="none",
                entity_reference_id=None,
                error_code=diag["code"],
                error_message=diag["message"],
                severity=severity,
                sheet_name=diag.get("sheet_name"),
                source_column=diag.get("source_column"),
                field_name=diag.get("field_name"),
                field_errors={
                    "recommended_correction": diag.get("recommended_correction"),
                    "related_entity": diag.get("related_entity"),
                    "source_row_number": diag.get("row_number"),
                },
                normalized_data={"normalized_value": diag.get("normalized_value")},
                row_data={"original_value": diag.get("original_value")},
                created_at=_now(),
                updated_at=_now(),
            )
        )
        next_row += 1

    batch.status = "validated" if blockers == 0 else "validation_failed"
    batch.valid_rows = batch.total_rows - blockers if batch.total_rows >= blockers else 0
    batch.invalid_rows = blockers
    batch.conflict_rows = 0
    batch.metadata_json = {**(batch.metadata_json or {}), "normalized_by_entity": normalized_by_entity}
    batch.updated_at = _now()

    await log_action(
        db=db,
        tenant_id=tenant.id,
        action="timetable_setup.workbook.validated",
        entity_type="ImportBatch",
        entity_id=batch.id,
        actor_id=actor.id,
        details={"blocker_count": blockers, "warning_count": warnings, "information_count": infos},
    )
    await db.commit()
    await db.refresh(batch)
    return {
        "batch": _batch_payload(batch),
        "diagnostics": {
            "blocker_count": blockers,
            "warning_count": warnings,
            "information_count": infos,
            "status": batch.status,
        },
    }


async def _resolve_scope(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    campus_code: str,
    academic_year_name: str,
    term_name: str,
) -> tuple[uuid.UUID, uuid.UUID, uuid.UUID] | None:
    campus = await db.scalar(select(Campus).where(Campus.tenant_id == tenant_id, func.lower(Campus.code) == _norm(campus_code), Campus.is_active.is_(True)))
    year = await db.scalar(select(AcademicYear).where(AcademicYear.tenant_id == tenant_id, func.lower(AcademicYear.name) == _norm(academic_year_name), AcademicYear.is_active.is_(True)))
    if year is None:
        return None
    term = await db.scalar(
        select(Term).where(
            Term.tenant_id == tenant_id,
            Term.academic_year_id == year.id,
            func.lower(Term.name) == _norm(term_name),
            Term.is_active.is_(True),
        )
    )
    if campus is None or term is None:
        return None
    return campus.id, year.id, term.id


@router.post("/workbooks/{batch_id}/commit", summary="Commit validated workbook transactionally")
async def commit_workbook(
    batch_id: uuid.UUID,
    tenant: Tenant = Depends(resolve_tenant),
    actor: User = Depends(resolve_authenticated_leadership),
    db: AsyncSession = Depends(get_db),
):
    _ensure_actor_tenant(actor, tenant)
    await set_tenant_context(db, tenant.id)

    batch = await _load_batch(db, tenant.id, batch_id, lock=True)
    if batch is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Workbook batch not found.")
    if batch.status == "committed":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Workbook already committed.")
    if batch.status != "validated":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Workbook must be validated before commit.")

    blocker_count = await db.scalar(
        select(func.count(ImportRowResult.id)).where(
            ImportRowResult.import_batch_id == batch.id,
            ImportRowResult.tenant_id == tenant.id,
            ImportRowResult.severity == "blocker",
        )
    )
    if int(blocker_count or 0) > 0:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Commit blocked by unresolved validation blockers.")

    normalized_by_entity = (batch.metadata_json or {}).get("normalized_by_entity", {})

    created = 0
    updated = 0
    unchanged = 0
    skipped = 0
    rejected = 0

    # Rooms
    for row in normalized_by_entity.get("rooms", []):
        room_code = _clean(row.get("room_code")).upper()
        room_name = _clean(row.get("room_name"))
        room_type = _clean(row.get("room_type"))
        if not room_code or not room_name or not room_type:
            rejected += 1
            continue
        campus_id = None
        if _clean(row.get("campus_code")):
            campus = await db.scalar(select(Campus).where(Campus.tenant_id == tenant.id, func.lower(Campus.code) == _norm(_clean(row.get("campus_code"))), Campus.is_active.is_(True)))
            if campus is None:
                rejected += 1
                continue
            campus_id = campus.id

        capacity = _parse_int(_clean(row.get("capacity"))) or 0
        capabilities = [item.strip() for item in _clean(row.get("specialist_capabilities")).split(",") if item.strip()]
        active_value = _parse_bool(_clean(row.get("active")))
        is_active = True if active_value is None else active_value

        existing = await db.scalar(select(TeachingRoom).where(TeachingRoom.tenant_id == tenant.id, TeachingRoom.room_code == room_code))
        if existing is None:
            db.add(
                TeachingRoom(
                    id=uuid.uuid4(),
                    tenant_id=tenant.id,
                    campus_id=campus_id,
                    room_code=room_code,
                    room_name=room_name,
                    room_type=room_type,
                    capacity=capacity,
                    floor_or_location=None,
                    specialist_capabilities=capabilities,
                    accessibility_notes=None,
                    source_type="excel_import",
                    review_status="approved",
                    is_active=is_active,
                )
            )
            created += 1
        else:
            changed = False
            if existing.room_name != room_name:
                existing.room_name = room_name
                changed = True
            if existing.room_type != room_type:
                existing.room_type = room_type
                changed = True
            if existing.capacity != capacity:
                existing.capacity = capacity
                changed = True
            if existing.specialist_capabilities != capabilities:
                existing.specialist_capabilities = capabilities
                changed = True
            if existing.is_active != is_active:
                existing.is_active = is_active
                changed = True
            existing.source_type = "excel_import"
            existing.review_status = "approved"
            if changed:
                updated += 1
            else:
                unchanged += 1

    # School week configs
    for row in normalized_by_entity.get("school_week", []):
        weekday = _parse_int(_clean(row.get("weekday")))
        operational = _parse_bool(_clean(row.get("operational")))
        if weekday is None or weekday < 0 or weekday > 6 or operational is None:
            rejected += 1
            continue
        campus_id = None
        if _clean(row.get("campus_code")):
            campus = await db.scalar(select(Campus).where(Campus.tenant_id == tenant.id, func.lower(Campus.code) == _norm(_clean(row.get("campus_code"))), Campus.is_active.is_(True)))
            if campus is None:
                rejected += 1
                continue
            campus_id = campus.id

        name = _clean(row.get("default_schedule_name")) or "Default Week"
        existing = await db.scalar(
            select(SchoolWeekConfig).where(
                SchoolWeekConfig.tenant_id == tenant.id,
                SchoolWeekConfig.campus_id == campus_id,
                SchoolWeekConfig.name == name,
                SchoolWeekConfig.is_active.is_(True),
            )
        )
        weekdays = [weekday] if operational else []
        if existing is None:
            db.add(
                SchoolWeekConfig(
                    id=uuid.uuid4(),
                    tenant_id=tenant.id,
                    campus_id=campus_id,
                    academic_year_id=None,
                    term_id=None,
                    name=name,
                    operational_weekdays=weekdays,
                    is_default=True,
                    source_type="excel_import",
                    review_status="approved",
                    is_active=True,
                )
            )
            created += 1
        else:
            if weekday not in existing.operational_weekdays and operational:
                existing.operational_weekdays = sorted(set(existing.operational_weekdays + [weekday]))
                existing.source_type = "excel_import"
                existing.review_status = "approved"
                updated += 1
            else:
                unchanged += 1

    # Bell schedules and periods
    schedule_by_name: dict[tuple[uuid.UUID | None, str], BellSchedule] = {}
    for row in normalized_by_entity.get("periods", []):
        schedule_name = _clean(row.get("schedule_name"))
        if not schedule_name:
            rejected += 1
            continue
        start_time = _clean(row.get("start_time"))
        end_time = _clean(row.get("end_time"))
        period_number = _parse_int(_clean(row.get("period_number")))
        if not TIME_RE.fullmatch(start_time) or not TIME_RE.fullmatch(end_time) or period_number is None or period_number <= 0:
            rejected += 1
            continue

        campus_id = None
        campus_code = _clean(row.get("campus_code"))
        if campus_code:
            campus = await db.scalar(select(Campus).where(Campus.tenant_id == tenant.id, func.lower(Campus.code) == _norm(campus_code), Campus.is_active.is_(True)))
            if campus is None:
                rejected += 1
                continue
            campus_id = campus.id

        key = (campus_id, schedule_name)
        schedule_item = schedule_by_name.get(key)
        if schedule_item is None:
            schedule_item = await db.scalar(
                select(BellSchedule).where(
                    BellSchedule.tenant_id == tenant.id,
                    BellSchedule.campus_id == campus_id,
                    BellSchedule.name == schedule_name,
                    BellSchedule.is_active.is_(True),
                )
            )
            if schedule_item is None:
                schedule_item = BellSchedule(
                    id=uuid.uuid4(),
                    tenant_id=tenant.id,
                    campus_id=campus_id,
                    academic_year_id=None,
                    term_id=None,
                    school_week_config_id=None,
                    name=schedule_name,
                    schedule_type="normal",
                    effective_start_date=None,
                    effective_end_date=None,
                    is_default=False,
                    source_type="excel_import",
                    review_status="approved",
                    is_active=True,
                    created_by_user_id=actor.id,
                    approved_by_user_id=actor.id,
                )
                db.add(schedule_item)
                created += 1
            schedule_by_name[key] = schedule_item

        label = _clean(row.get("period_label")) or f"Period {period_number}"
        teaching_period = _parse_bool(_clean(row.get("teaching_period")))
        is_teaching = True if teaching_period is None else teaching_period
        period = await db.scalar(
            select(BellSchedulePeriod).where(
                BellSchedulePeriod.tenant_id == tenant.id,
                BellSchedulePeriod.bell_schedule_id == schedule_item.id,
                BellSchedulePeriod.period_number == period_number,
                BellSchedulePeriod.is_active.is_(True),
            )
        )
        if period is None:
            db.add(
                BellSchedulePeriod(
                    id=uuid.uuid4(),
                    tenant_id=tenant.id,
                    bell_schedule_id=schedule_item.id,
                    applicable_grade_level_id=None,
                    period_number=period_number,
                    label=label,
                    start_time=start_time,
                    end_time=end_time,
                    is_teaching_period=is_teaching,
                    is_break=not is_teaching and _norm(_clean(row.get("period_type"))) == "break",
                    is_lunch=not is_teaching and _norm(_clean(row.get("period_type"))) == "lunch",
                    is_active=True,
                )
            )
            created += 1
        else:
            changed = False
            if period.label != label:
                period.label = label
                changed = True
            if period.start_time != start_time:
                period.start_time = start_time
                changed = True
            if period.end_time != end_time:
                period.end_time = end_time
                changed = True
            if period.is_teaching_period != is_teaching:
                period.is_teaching_period = is_teaching
                changed = True
            if changed:
                updated += 1
            else:
                unchanged += 1

    # Weekly teaching requirements
    for row in normalized_by_entity.get("teaching_requirements", []):
        scope = await _resolve_scope(
            db,
            tenant.id,
            _clean(row.get("campus_code")),
            _clean(row.get("academic_year")),
            _clean(row.get("term")),
        )
        if scope is None:
            rejected += 1
            continue
        campus_id, year_id, term_id = scope

        class_code = _clean(row.get("class_code"))
        subject_code = _clean(row.get("subject_code"))
        class_item = await db.scalar(select(Class).where(Class.tenant_id == tenant.id, func.lower(Class.code) == _norm(class_code), Class.is_active.is_(True)))
        subject_item = await db.scalar(select(Subject).where(Subject.tenant_id == tenant.id, func.lower(Subject.code) == _norm(subject_code)))
        if class_item is None or subject_item is None:
            rejected += 1
            continue

        teacher_id_value = _clean(row.get("teacher_id"))
        teacher_item = None
        if teacher_id_value:
            teacher_item = await db.scalar(select(Teacher).where(Teacher.tenant_id == tenant.id, func.lower(Teacher.employee_id) == _norm(teacher_id_value)))

        sessions_per_week = _parse_int(_clean(row.get("sessions_per_week")))
        periods_per_session = _parse_int(_clean(row.get("periods_per_session")) or "1")
        min_daily = _parse_int(_clean(row.get("minimum_daily_sessions")) or "0")
        max_daily = _parse_int(_clean(row.get("maximum_daily_sessions")) or "3")
        priority = _parse_int(_clean(row.get("priority")) or "100")
        if sessions_per_week is None or sessions_per_week <= 0 or periods_per_session is None or periods_per_session <= 0:
            rejected += 1
            continue
        if min_daily is None or max_daily is None or max_daily < min_daily:
            rejected += 1
            continue

        preferred_periods = _parse_period_list(_clean(row.get("preferred_periods")))
        forbidden_periods = _parse_period_list(_clean(row.get("forbidden_periods")))
        double_mode = _norm(_clean(row.get("double_period_preference")))
        if double_mode not in {"none", "preferred", "required"}:
            double_mode = "none"

        existing = await db.scalar(
            select(WeeklyTeachingRequirement).where(
                WeeklyTeachingRequirement.tenant_id == tenant.id,
                WeeklyTeachingRequirement.campus_id == campus_id,
                WeeklyTeachingRequirement.academic_year_id == year_id,
                WeeklyTeachingRequirement.term_id == term_id,
                WeeklyTeachingRequirement.class_id == class_item.id,
                WeeklyTeachingRequirement.subject_id == subject_item.id,
                WeeklyTeachingRequirement.is_active.is_(True),
            )
        )

        if existing is None:
            db.add(
                WeeklyTeachingRequirement(
                    id=uuid.uuid4(),
                    tenant_id=tenant.id,
                    campus_id=campus_id,
                    academic_year_id=year_id,
                    term_id=term_id,
                    class_id=class_item.id,
                    subject_id=subject_item.id,
                    teacher_id=teacher_item.id if teacher_item else None,
                    sessions_per_week=sessions_per_week,
                    periods_per_session=periods_per_session,
                    min_daily_sessions=min_daily,
                    max_daily_sessions=max_daily,
                    double_period_mode=double_mode,
                    specialist_room_type=_clean(row.get("required_room_type")) or None,
                    preferred_period_numbers=preferred_periods,
                    forbidden_period_numbers=forbidden_periods,
                    has_fixed_sessions=False,
                    fixed_session_rules=[],
                    priority=priority or 100,
                    source_type="excel_import",
                    review_status="approved",
                    is_active=True,
                )
            )
            created += 1
        else:
            changed = False
            for attr, value in (
                ("teacher_id", teacher_item.id if teacher_item else None),
                ("sessions_per_week", sessions_per_week),
                ("periods_per_session", periods_per_session),
                ("min_daily_sessions", min_daily),
                ("max_daily_sessions", max_daily),
                ("double_period_mode", double_mode),
                ("specialist_room_type", _clean(row.get("required_room_type")) or None),
                ("preferred_period_numbers", preferred_periods),
                ("forbidden_period_numbers", forbidden_periods),
                ("priority", priority or 100),
            ):
                if getattr(existing, attr) != value:
                    setattr(existing, attr, value)
                    changed = True
            existing.source_type = "excel_import"
            existing.review_status = "approved"
            if changed:
                updated += 1
            else:
                unchanged += 1

    batch.status = "committed"
    batch.created_rows = created
    batch.updated_rows = updated
    batch.skipped_rows = skipped + unchanged
    batch.conflict_rows = rejected
    batch.committed_at = _now()
    batch.completed_at = _now()

    readiness = await compute_timetable_input_readiness(db, tenant.id)

    await log_action(
        db=db,
        tenant_id=tenant.id,
        action="timetable_setup.workbook.committed",
        entity_type="ImportBatch",
        entity_id=batch.id,
        actor_id=actor.id,
        details={
            "created": created,
            "updated": updated,
            "unchanged": unchanged,
            "skipped": skipped,
            "rejected": rejected,
            "readiness_blockers": readiness.get("blocker_count", 0),
        },
    )

    await db.commit()
    await db.refresh(batch)
    return {
        "batch": _batch_payload(batch),
        "commit_summary": {
            "create": created,
            "update": updated,
            "unchanged": unchanged,
            "skipped": skipped,
            "rejected": rejected,
        },
        "readiness": {
            "blocker_count": readiness.get("blocker_count", 0),
            "warning_count": readiness.get("warning_count", 0),
            "information_count": readiness.get("information_count", 0),
            "is_generation_ready": readiness.get("is_generation_ready", False),
        },
    }


@router.post("/workbooks/{batch_id}/cancel", summary="Cancel workbook batch")
async def cancel_workbook_batch(
    batch_id: uuid.UUID,
    tenant: Tenant = Depends(resolve_tenant),
    actor: User = Depends(resolve_authenticated_leadership),
    db: AsyncSession = Depends(get_db),
):
    _ensure_actor_tenant(actor, tenant)
    await set_tenant_context(db, tenant.id)
    batch = await _load_batch(db, tenant.id, batch_id, lock=True)
    if batch is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Workbook batch not found.")
    if batch.status == "committed":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Committed batches cannot be cancelled.")
    if batch.status == "cancelled":
        return _batch_payload(batch)

    batch.status = "cancelled"
    batch.completed_at = _now()
    await log_action(
        db=db,
        tenant_id=tenant.id,
        action="timetable_setup.workbook.cancelled",
        entity_type="ImportBatch",
        entity_id=batch.id,
        actor_id=actor.id,
        details={},
    )
    await db.commit()
    await db.refresh(batch)
    return _batch_payload(batch)


@router.get("/workbooks/{batch_id}/diagnostics", summary="Get paginated workbook diagnostics")
async def get_workbook_diagnostics(
    batch_id: uuid.UUID,
    severity: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=200),
    tenant: Tenant = Depends(resolve_tenant),
    actor: User = Depends(resolve_authenticated_leadership),
    db: AsyncSession = Depends(get_db),
):
    _ensure_actor_tenant(actor, tenant)
    await set_tenant_context(db, tenant.id)
    batch = await _load_batch(db, tenant.id, batch_id)
    if batch is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Workbook batch not found.")

    stmt = select(ImportRowResult).where(
        ImportRowResult.import_batch_id == batch.id,
        ImportRowResult.tenant_id == tenant.id,
        ImportRowResult.severity.is_not(None),
    ).order_by(ImportRowResult.row_number.asc())
    if severity:
        stmt = stmt.where(ImportRowResult.severity == severity)

    rows = (await db.execute(stmt)).scalars().all()
    items = [
        {
            "sheet": row.sheet_name,
            "row_number": (row.field_errors or {}).get("source_row_number"),
            "source_column": row.source_column,
            "field": row.field_name,
            "code": row.error_code,
            "severity": row.severity,
            "message": row.error_message,
            "original_value": (row.row_data or {}).get("original_value"),
            "normalized_value": (row.normalized_data or {}).get("normalized_value"),
            "related_entity": (row.field_errors or {}).get("related_entity"),
            "recommended_correction": (row.field_errors or {}).get("recommended_correction"),
        }
        for row in rows
    ]
    return _list_to_page(items, page, page_size)
